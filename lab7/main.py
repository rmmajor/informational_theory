import math
import random

from Tools.demo.spreadsheet import Sheet
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border


# перевірка на парність
def encode_with_parity_check(codebook: dict[str | int, str]) -> dict[str | int, str]:
    encoded_codebook = {}

    for key in codebook:
        code_word = codebook[key]

        parity_bit = '1' if code_word.count('1') % 2 == 1 else '0'
        encoded_code_word = code_word + parity_bit

        encoded_codebook[key] = encoded_code_word

    return encoded_codebook


def parity_check_mistakes_detection(encoded_codebook: dict[str | int, str]) -> dict[str | int, tuple[str, bool]]:
    mistakes: dict[str | int, tuple[str, bool]] = {}
    for key in encoded_codebook:
        full_code_word = encoded_codebook[key]
        code_word = encoded_codebook[key][:-1]
        parity_bit = encoded_codebook[key][-1]
        if (code_word.count('1') % 2) != int(parity_bit):
            mistakes[key] = (full_code_word, True)
        else:
            mistakes[key] = (full_code_word, False)

    return mistakes


# інверсний код Бауера
def encode_with_bauer(codebook: dict[str | int, str]) -> dict[str | int, str]:
    encoded_codebook = {}

    for key in codebook:
        code_word = codebook[key]

        if code_word.count('1') % 2 == 1:
            check_word = ''
            for c in code_word:
                check_word += '1' if c == '0' else '0'
        else:
            check_word = code_word

        encoded_code_word = code_word + check_word
        encoded_codebook[key] = encoded_code_word

    return encoded_codebook


def bauer_mistakes_detection(encoded_codebook: dict[str | int, str]) -> dict[str | int, tuple[str, bool]]:
    mistakes: dict[str | int, tuple[str, bool]] = {}
    for key in encoded_codebook:
        full_code_word = encoded_codebook[key]
        k = len(full_code_word) // 2
        code_word = full_code_word[:k]
        check_word = full_code_word[k:]

        if code_word.count('1') % 2 == 1:
            inversed_code_word = ''
            for c in code_word:
                inversed_code_word += '1' if c == '0' else '0'
            mistakes[key] = (full_code_word, inversed_code_word != check_word)
        else:
            mistakes[key] = (full_code_word, code_word != check_word)

    return mistakes


# кореляційний код
def encode_with_correlational_code(codebook: dict[str | int, str]) -> dict[str | int, str]:
    encoded_codebook: dict[str | int, str] = {}

    for key in codebook:
        code_word = codebook[key]
        encoded_code_word = ''
        for bit in code_word:
            if bit == '0':
                encoded_code_word += '01'
            else:
                encoded_code_word += '10'
        encoded_codebook[key] = encoded_code_word

    return encoded_codebook


def correlational_code_mistakes_detection(encoded_codebook: dict[str | int, str]) -> dict[str | int, tuple[str, bool]]:
    mistakes: dict[str | int, tuple[str, bool]] = {}
    for key in encoded_codebook:
        code_word = encoded_codebook[key]
        is_mistake = False
        for i in range(0, len(code_word), 2):
            if code_word[i] == code_word[i+1]:
                is_mistake = True
                break

        mistakes[key] = (code_word, is_mistake)
    return mistakes


# код Бергера
def calculate_berger_check_bits_number(size: int) -> int:
     return math.ceil(math.log2(size + 1))


def calculate_number_of_ones_in_code_word(code_word: str) -> int:
    number_of_ones = 0
    for bit in code_word:
        number_of_ones += (bit == '1')

    return number_of_ones


def int_to_bin(num: int) -> str:
    return bin(num)[2:]      # from '0b101010' to '101010'


def add_leading_zeros_if_needed(bin_number_of_ones: str, check_bits_number: int) -> str:
    number_of_leading_zeros_to_add = check_bits_number - len(bin_number_of_ones)
    return ('0' * number_of_leading_zeros_to_add) + bin_number_of_ones


def inverse_bin_number(bin_number: str) -> str:
    inversed_bin_number: str = ''

    for bit in bin_number:
        inversed_bin_number += ('1' if bit == '0' else '0')

    return inversed_bin_number


def calculate_check_bits(code_word) -> str:
    check_bits_number: int = calculate_berger_check_bits_number(size=len(code_word))
    number_of_ones: int = calculate_number_of_ones_in_code_word(code_word)
    bin_number_of_ones: str = int_to_bin(number_of_ones)
    bin_number_of_ones_with_leading_zeros: str = add_leading_zeros_if_needed(bin_number_of_ones, check_bits_number)
    return inverse_bin_number(bin_number_of_ones_with_leading_zeros)


def encode_single_word_with_berger(code_word: str) -> str:
    check_bits: str = calculate_check_bits(code_word)
    return code_word + check_bits


def encode_with_berger(codebook: dict[str | int, str]) -> dict[str | int, str]:
    encoded_codebook: dict[str | int, str] = {}

    for key in codebook:
        encoded_codebook[key] = encode_single_word_with_berger(codebook[key])

    return encoded_codebook


def calculate_number_of_check_bits_for_encoded_word(encoded_word: str):
    n = len(encoded_word)
    r = 0
    for k in range(1, n+1):
        r = calculate_berger_check_bits_number(k)
        if r + k == n:
            break

    return r


def berger_detect_mistake_in_single_word(encoded_word: str) -> tuple[str, bool]:
    number_of_check_bits = calculate_number_of_check_bits_for_encoded_word(encoded_word)
    initial_check_bits = encoded_word[-number_of_check_bits:]
    recalculated_check_bits = calculate_check_bits(code_word=encoded_word[:-number_of_check_bits])
    # print(number_of_check_bits, initial_check_bits, recalculated_check_bits, encoded_word[:-number_of_check_bits])

    return encoded_word, (initial_check_bits != recalculated_check_bits)


def berger_mistakes_detection(encoded_codebook: dict[str | int, str]) -> dict[str | int, tuple[str, bool]]:
    mistakes: dict[str | int, tuple[str, bool]] = {}
    for key in encoded_codebook:
        mistakes[key] = berger_detect_mistake_in_single_word(encoded_codebook[key])

    return mistakes


# код Хеммінга
def powers_of_two(i: int) -> list[int]:
    powers: list[int] = []
    power: int = 0
    while i > 0:
        if i % 2 == 1:
            powers.append(2 ** power)
        power += 1
        i //= 2
    powers.reverse()

    return powers


def calculate_number_of_hamming_check_bits(code_word_size: int) -> int:
    check_bits_number = 0
    while 2 ** check_bits_number < code_word_size + check_bits_number + 1:
        check_bits_number += 1

    return check_bits_number


def calculate_check_bits_in_code_list_with_check_bits(code_bits: list[int]) -> list[int]:
    for i in range(len(code_bits)):
        powers: list[int] = powers_of_two(i+1)
        if len(powers) == 1:
            continue

        for power in powers:
            code_bits[power-1] ^= code_bits[i]

    return code_bits


def make_bit_list_with_check_bits_set_to_zero_from_string(code_word: str) -> list[int]:
    r = calculate_number_of_hamming_check_bits(len(code_word))
    encoded_bits: list[int] = [0] * (len(code_word) + r)
    j = 0
    for i in range(len(encoded_bits)):
        if i + 1 not in [2 ** k for k in range(r)]:
            encoded_bits[i] = int(code_word[j])
            j += 1

    return encoded_bits


def hamming_encode_single_word(code_word: str) -> str:
    code_word_with_zeroed_check_bits: list[int] = make_bit_list_with_check_bits_set_to_zero_from_string(code_word)
    encoded_bits = calculate_check_bits_in_code_list_with_check_bits(code_word_with_zeroed_check_bits)

    # перетворюєм ліст бітів на строку і вертаєм
    return ''.join([str(bit) for bit in encoded_bits])


def encode_with_hamming(codebook: dict[str | int, str]) -> dict[str | int, str]:
    encoded_codebook: dict[str | int, str] = {}

    for key in codebook:
        encoded_codebook[key] = hamming_encode_single_word(codebook[key])

    return encoded_codebook


def calculate_check_bits_within_hamming_code(size: int) -> int:
    check_bits_number = 0
    while 2 ** check_bits_number < size:
        check_bits_number += 1

    return check_bits_number


def check_if_code_has_mistake(encoded_code: str, recalculated_code: str) -> bool:
    return encoded_code != recalculated_code


def locate_mistaken_bit(encoded_code: str, recalculated_code: str) -> int:
    k = calculate_check_bits_within_hamming_code(len(encoded_code))
    check_bits_positions = [2 ** i - 1 for i in range(k)]
    mistake_index = 0

    for pos in check_bits_positions:
        if encoded_code[pos] != recalculated_code[pos]:
            mistake_index += pos + 1

    return mistake_index


def create_hamming_bits_list_from_encoded_string(encoded_code: str) -> list[int]:
    encoded_bits_list: list[int] = [int(x) for x in encoded_code]
    k = calculate_check_bits_within_hamming_code(len(encoded_code))
    check_bits_positions = [2**i-1 for i in range(k)]
    for pos in check_bits_positions:
        encoded_bits_list[pos] = 0

    return encoded_bits_list


def hamming_single_word_mistake_detection(encoded_code: str) -> tuple[str, bool | int]:
    """if code has mistake, returns index of wrong bit, otherwise, returns False"""

    encoded_bits_list: list[int] = create_hamming_bits_list_from_encoded_string(encoded_code)
    encoded_bits_with_recalculated_check_bits: list[int] = calculate_check_bits_in_code_list_with_check_bits(
        encoded_bits_list
    )
    encoded_code_with_recalculated_check_bits = \
        ''.join([str(bit) for bit in encoded_bits_with_recalculated_check_bits])

    is_mistake = check_if_code_has_mistake(encoded_code, encoded_code_with_recalculated_check_bits)
    if is_mistake:
        mistake_index = locate_mistaken_bit(encoded_code, encoded_code_with_recalculated_check_bits)
        return encoded_code, mistake_index

    return encoded_code, False


def hamming_mistakes_detection(encoded_codebook: dict[str | int, str]) -> dict[str | int, tuple[str, bool]]:
    mistakes: dict[str | int, tuple[str, bool]] = {}
    for key in encoded_codebook:
        mistakes[key] = hamming_single_word_mistake_detection(encoded_codebook[key])

    return mistakes


# mistakes creation
def make_mistake_in_code(code: str) -> str:
    random_bit_index: int = random.randint(0, len(code) - 1)
    random_bit: str = '1' if code[random_bit_index] == '0' else '0'
    new_code: str = code[:random_bit_index] + random_bit + code[random_bit_index + 1:]
    return new_code


def make_mistakes_in_codebook(codes: dict[str | int, str]) -> dict[str, str]:
    mistaken_codes = codes.copy()
    amount_of_words_to_make_mistake: int = random.randint(4, len(mistaken_codes)-3)
    keys_to_make_mistake: list[str | int] = random.sample(list(mistaken_codes.keys()), amount_of_words_to_make_mistake)
    print("Слова в чиїх кодах буде помилка:", keys_to_make_mistake)

    for key in keys_to_make_mistake:
        mistaken_codes[key] = make_mistake_in_code(mistaken_codes[key])

    return mistaken_codes


def read_lines():
    lines = []
    while True:
        line = input()
        if not line:
            break
        lines.append(line)
    return lines


def read_codebook():
    print('Введіть символи')
    words = read_lines()
    print('Введіть коди')
    codes = read_lines()
    codebook = {}

    for word, code in zip(words, codes):
        codebook[word] = code

    return codebook


def draw_table(code_words, encoded_code_words, encoded_code_words_with_mistakes, is_there_mistake):
    table_data = []
    print(code_words.keys())
    for key in code_words.keys():

        if is_there_mistake[key][1] > 0 and not isinstance(is_there_mistake[key][1], bool):
            mistake = '<- ' + str(is_there_mistake[key][1])
        else:
            mistake = '<-' if is_there_mistake[key][1] else ''

        table_data.append([code_words[key],
                           encoded_code_words[key],
                           encoded_code_words_with_mistakes[key],
                           mistake
                           ])

    headers = ["Кодове слово", "Закодовані кодові слова",
               "Закодовані кодові слова з помилкою", "Виділені кодові слова або біти з помилкою"]

    table = tabulate(table_data, headers)
    print(table)


def write_in_exel_table(title: str, code_words, encoded_code_words, encoded_code_words_with_mistakes, is_there_mistake):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лабораторна робота №7"
    ws['A1'] = "Кодове слово"
    ws['B1'] = "Закодовані кодові слова"
    ws['C1'] = "Закодовані кодові слова з помилкою"
    ws['D1'] = "Виділені кодові слова або біти з помилкою"
    row = 2
    for key in code_words.keys():
        ws[f'A{row}'] = code_words[key]
        ws[f'B{row}'] = encoded_code_words[key]
        ws[f'C{row}'] = encoded_code_words_with_mistakes[key]
        ws[f'D{row}'] = is_there_mistake[key][1]
        row += 1

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    # add borders
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=row-1):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(f"{title}.xlsx")


def do_lab(codes):
    print("Початковий кодовий словник:", codes)

    # перевірка на парність
    print()
    encoded_with_parity_check = encode_with_parity_check(codes)
    print("Кодовий словник, який закодовано перевіркою на парність", encoded_with_parity_check)
    encoded_with_parity_check_with_mistakes = make_mistakes_in_codebook(encoded_with_parity_check)
    print("Закодовані кодові слова з помилками", encoded_with_parity_check_with_mistakes)
    parity_check_mistakes = parity_check_mistakes_detection(encoded_with_parity_check_with_mistakes)
    print("Виділені кодові слова або біти з помилкою", parity_check_mistakes)
    draw_table(codes, encoded_with_parity_check, encoded_with_parity_check_with_mistakes, parity_check_mistakes)
    write_in_exel_table("Перевірка на парність",
                        codes,
                        encoded_with_parity_check,
                        encoded_with_parity_check_with_mistakes,
                        parity_check_mistakes)

    # інверсний код Бауера
    print()
    encoded_with_bauer = encode_with_bauer(codes)
    print("Кодовий словник, який закодовано інверсним кодом Бауера", encoded_with_bauer)
    encoded_with_bauer_with_mistakes = make_mistakes_in_codebook(encoded_with_bauer)
    print("Закодовані кодові слова з помилками", encoded_with_bauer_with_mistakes)
    bauer_mistakes = bauer_mistakes_detection(encoded_with_bauer_with_mistakes)
    print("Виділені кодові слова або біти з помилкою", bauer_mistakes)
    draw_table(codes, encoded_with_bauer, encoded_with_bauer_with_mistakes, bauer_mistakes)
    write_in_exel_table("Інверсний код Бауера",
                        codes,
                        encoded_with_bauer,
                        encoded_with_bauer_with_mistakes,
                        bauer_mistakes)

    # кореляційний код
    print()
    encoded_with_correlational_code = encode_with_correlational_code(codes)
    print("Кодовий словник, який закодовано кореляційним кодом", encoded_with_correlational_code)
    encoded_with_correlational_code_with_mistakes = make_mistakes_in_codebook(encoded_with_correlational_code)
    print("Закодовані кодові слова з помилками", encoded_with_correlational_code_with_mistakes)
    correlational_code_mistakes = correlational_code_mistakes_detection(encoded_with_correlational_code_with_mistakes)
    print("Виділені кодові слова або біти з помилкою", correlational_code_mistakes)
    draw_table(codes, encoded_with_correlational_code,
               encoded_with_correlational_code_with_mistakes, correlational_code_mistakes)
    write_in_exel_table("Кореляційний код",
                        codes,
                        encoded_with_correlational_code,
                        encoded_with_correlational_code_with_mistakes,
                        correlational_code_mistakes)

    # код Бергера
    print()
    encoded_with_berger = encode_with_berger(codes)
    print("Кодовий словник, який закодовано кодом, Бергера", encoded_with_berger)
    encoded_with_berger_with_mistakes = make_mistakes_in_codebook(encoded_with_berger)
    print("Закодовані кодові слова з помилками", encoded_with_berger_with_mistakes)
    berger_mistakes = berger_mistakes_detection(encoded_with_berger_with_mistakes)
    print("Виділені кодові слова або біти з помилкою", berger_mistakes)
    draw_table(codes, encoded_with_berger, encoded_with_berger_with_mistakes, berger_mistakes)
    write_in_exel_table("Код Бергера",
                        codes,
                        encoded_with_berger,
                        encoded_with_berger_with_mistakes,
                        berger_mistakes)

    # код Хеммінга
    print()
    encoded_with_hamming = encode_with_hamming(codes)
    print("Кодовий словник, який закодовано кодом Хеммінга", encoded_with_hamming)
    encoded_with_hamming_with_mistakes = make_mistakes_in_codebook(encoded_with_hamming)
    print("Закодовані кодові слова з помилками", encoded_with_hamming_with_mistakes)
    hamming_code_mistakes = hamming_mistakes_detection(encoded_with_hamming_with_mistakes)
    print("Виділені кодові слова або біти з помилкою", hamming_code_mistakes)
    draw_table(codes, encoded_with_hamming, encoded_with_hamming_with_mistakes, hamming_code_mistakes)
    write_in_exel_table("Код Хеммінга",
                        codes,
                        encoded_with_hamming,
                        encoded_with_hamming_with_mistakes,
                        hamming_code_mistakes)


def main():
    print('='*20)
    print('Введіть кодову множину')
    codes = read_codebook()
    do_lab(codes)


if __name__ == '__main__':
    main()
